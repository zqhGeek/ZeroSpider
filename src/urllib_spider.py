# coding :utf8
import os
import time
from urllib import request
from urllib import error
import chardet
from bs4 import BeautifulSoup
from urllib.request import urlretrieve
import openpyxl


def proxy_http_connect():
    # 这是代理IP
    url = "http://www.baidu.com"
    # 代理地址
    proxy = {'http': '124.205.155.151:9090'}
    # 创建ProxyHandler
    proxy_support = request.ProxyHandler(proxy)
    # 创建Opener
    opener = request.build_opener(proxy_support)
    # 添加User Angent
    opener.addheaders = [('User-Agent',
                          'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36')]
    # 安装OPener
    request.install_opener(opener)
    try:
        # response = opener.open(url)
        response = request.urlopen(url)
        print("geturl打印信息：%s" % (response.geturl()))
        print('**********************************************')
        print("info打印信息：%s" % (response.info()))
        print('**********************************************')
        print("getcode打印信息：%s" % (response.getcode()))
        html = response.read()
        html = html.decode("utf-8")
        print(html)
    except error.URLError as e:
        if type(e) == error.URLError:
            print("URL错误" + str(e))
        if type(e) == error.HTTPError:
            print("HTTP错误" + str(e))


def http_connect(http_address="http://www.baidu.com"):
    req = request.Request(http_address)
    # 传入headers
    req.add_header('User-Agent',
                   'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36')
    try:
        response = request.urlopen(req)
        print("geturl打印信息：%s" % (response.geturl()))
        print('**********************************************')
        print("info打印信息：%s" % (response.info()))
        print('**********************************************')
        print("getcode打印信息：%s" % (response.getcode()))
        print('**********************************************')
        html = response.read()
        charset = chardet.detect(html)
        print("charset编码格式：%s" % (charset["encoding"]))
        print('**********************************************')
        if charset["encoding"].find("2312") > 0:
            html = html.decode("gb18030")
        else:
            html = html.decode(charset["encoding"])
        return html
    except error.URLError as e:
        if type(e) == error.URLError:
            print("URL错误" + str(e))
        if type(e) == error.HTTPError:
            print("HTTP错误" + str(e))
        return None


# def cookie_login():

def beautifulsoup_spider():
    download_html = http_connect("http://www.biqukan.com/1_1094/5403177.html")
    soup_texts = BeautifulSoup(download_html, 'lxml')
    texts = soup_texts.find_all(id='content', class_='showtxt')
    soup_text = BeautifulSoup(str(texts), 'lxml')
    # 将\xa0无法解码的字符删除
    return soup_text.div.text


def save_to_excel(age, school, job_address, home_address, money, job, marry_time):
    if "统计表格.xlsx" not in os.listdir("D:/IDEA_Project/收集器"):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "信息表"
        # sheet = wb.create_sheet(u"信息表")
        TITLE = ['出生年月', '学历', '就职地', '户籍地', '年薪', '职业', '计划结婚']
        DATA = [age, school, job_address, home_address, money, job, marry_time]
        sheet.append(TITLE)
        sheet.append(DATA)
        wb.save("D:/IDEA_Project/收集器/统计表格.xlsx")
    else:
        wb = openpyxl.load_workbook("D:/IDEA_Project/收集器/统计表格.xlsx")
        sheet = wb.get_sheet_by_name('信息表')
        DATA = [age, school, job_address, home_address, money, job, marry_time]
        sheet.append(DATA)
        wb.save("D:/IDEA_Project/收集器/统计表格.xlsx")


def down_girl(path, girl_id):
    try:
        print("NO:" + girl_id)
        downloadHtml = http_connect("http://date.jobbole.com/" + girl_id + "/")
        soupTexts = BeautifulSoup(downloadHtml, 'lxml')
        textsFind = soupTexts.find(class_="p-entry")
        textsTitle = soupTexts.find(class_="p-tit-single")
        print("标题:" + textsTitle.text)
        if str(textsTitle.text).count("脱单") > 0 or str(textsTitle.text).count("【") > 0:
            print("已脱单")
            return False
        filePath = path + "/" + str(textsTitle.text)
        if str(textsTitle.text) not in os.listdir(path):
            os.makedirs(filePath)
        else:
            print("已存在")
            return True
        tagFind = BeautifulSoup(str(textsFind), 'lxml')
        girlAttribute = str(tagFind.p).replace("\n", "").replace("<p>", "").replace("</p>", "").replace("<br/>", "\n")
        attributeList = girlAttribute.split("\n")
        if len(attributeList) < 4 or str(attributeList[0]).find("出生年月") < 0:
            print("已脱单")
            return False
        with open(filePath + "/" + "信息卡.txt", 'w', encoding='utf-8', newline='') as f:
            f.write(girlAttribute)
        img_list = textsFind.find_all("img")
        i = 1
        for imgs in img_list:
            imgs_str = str(imgs)
            urlretrieve(imgs_str[imgs_str.index("http"):imgs_str.index("jpg") + 3], filePath + "/" + str(i) + ".jpg")
            i += 1
        print("**********************************************")
        print(attributeList[0])
        print(attributeList[3])
        print(attributeList[4])
        print(attributeList[7])
        save_to_excel(attributeList[0][5:], attributeList[3][3:], attributeList[4][5:], attributeList[5][3:], attributeList[8][5:],
                      attributeList[7][3:], attributeList[11][8:])
        # if str(attributeList[-2]).find("http") > 0:
        #     urlretrieve(str(attributeList[-2])[
        #                 attributeList[-2].index("http"):attributeList[-2].index("jpg") + 3], filePath + "/" + "01.jpg")
        # if str(attributeList[-1]).find("http") > 0:
        #     urlretrieve(str(attributeList[-1])[
        #                 attributeList[-1].index("http"):attributeList[-1].index("jpg") + 3], filePath + "/" + "02.jpg")
        return True
    except Exception:
        print("错误")
        return False


class RetryModel:
    def __init__(self, retry_count, retry_time, retry_id):
        self.count = retry_count
        self.time = retry_time
        self.id = retry_id

    def update(self, retry_time):
        self.count += 1
        self.time = retry_time

    def should_retry(self):
        if self.count < 3:
            return True
        else:
            return False

    def get_count(self):
        return self.count

    def get_id(self):
        return self.id


if __name__ == "__main__":
    if "收集器" not in os.listdir("D:/IDEA_Project"):
        os.makedirs("D:/IDEA_Project/收集器")
        pass
    page = 1
    isNext = True
    retryList = []
    while isNext:
        isHttp = False
        while not isHttp:
            # 只要失败就重试
            try:
                download_html = http_connect("http://date.jobbole.com/page/" + str(page))
                soup_texts = BeautifulSoup(download_html, 'lxml')
                isHttp = True
            except:
                print("获取列表失败，重试")
                time.sleep(2)
        print("页数：" + str(page))
        numberList = []
        if len(soup_texts.find_all(class_="media")) > 0:
            page += 1
            for media in soup_texts.find_all(class_="media"):
                try:
                    id_ = BeautifulSoup(str(media), 'lxml').html.body.li.a["data-post-id"]
                    numberList.append(id_)
                    print(id_)
                except KeyError:
                    pass
            for girl in numberList:
                if not down_girl("D:/IDEA_Project/收集器", girl):
                    retryList.append(RetryModel(1, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), girl))
                time.sleep(2)

        else:
            if len(retryList) > 0:
                for retryModel in retryList:
                    while retryModel.should_retry():
                        if not down_girl("D:/IDEA_Project/收集器", retryModel.get_id()):
                            retryModel.update(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
                        else:
                            break
                        time.sleep(2)
                isNext = False
            else:
                isNext = False
